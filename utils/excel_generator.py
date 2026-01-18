import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import logging

logger = logging.getLogger(__name__)

def generate_excel(data: dict, output_path: str) -> str:
    """
    Genera un archivo Excel con los datos del documento.
    
    Args:
        data: Diccionario con los datos del documento
        output_path: Ruta donde guardar el Excel
    
    Returns:
        Ruta del archivo Excel generado
    """
    try:
        # Crear libro de trabajo
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos Documento"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = ["CAMPO", "VALOR", "DESCRIPCIÓN"]
        ws.append(headers)
        
        # Aplicar estilos a encabezados
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Datos del documento
        datos = [
            ("NÚMERO DE PROCESO", data.get("process_number", "N/A"), "Identificador único del proceso"),
            ("PROVEEDOR", data.get("provider", "N/A"), "Nombre del proveedor o empresa"),
            ("CONDUCTOR", data.get("driver", "N/A"), "Nombre del conductor"),
            ("PLACA TRACTO", data.get("plate_tractor", "N/A"), "Placa del vehículo tracto"),
            ("PLACA REMOLQUE", data.get("plate_remolque", "N/A"), "Placa del remolque"),
            ("PESO NETO", f"{data.get('net_weight', 0):,.2f} kg", "Peso neto de la carga"),
            ("PESO BRUTO", data.get("peso_bruto", "N/A"), "Peso bruto total"),
            ("TARA", data.get("tara", "N/A"), "Peso de la tara"),
            ("PRODUCTO", data.get("product", "N/A"), "Tipo de producto/material"),
            ("CANTIDAD", data.get("cantidad", "N/A"), "Cantidad de unidades"),
            ("UNIDAD", data.get("unidad_medida", "N/A"), "Unidad de medida"),
            ("HUMEDAD", f"{data.get('humedad', 0)}%", "Porcentaje de humedad"),
            ("IMPUREZAS", f"{data.get('impurezas', 0)}%", "Porcentaje de impurezas"),
            ("TEMPERATURA", f"{data.get('temperatura', 0)}°C", "Temperatura del producto"),
            ("ESTADO", data.get("estado", "N/A"), "Estado de aprobación"),
            ("FECHA RECEPCIÓN", data.get("fecha", "N/A"), "Fecha de recepción"),
            ("HORA RECEPCIÓN", data.get("hora", "N/A"), "Hora de recepción"),
            ("TRANSPORTADORA", data.get("transportadora", "N/A"), "Empresa transportadora"),
            ("NIT PROVEEDOR", data.get("nit_proveedor", "N/A"), "NIT del proveedor"),
            ("CÉDULA CONDUCTOR", data.get("cedula", "N/A"), "Cédula del conductor"),
        ]
        
        # Agregar datos a la hoja
        for row_data in datos:
            ws.append(row_data)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Agregar hoja de observaciones
        ws2 = wb.create_sheet(title="Observaciones")
        ws2.append(["OBSERVACIONES DEL DOCUMENTO"])
        ws2['A1'].font = Font(bold=True, size=14)
        
        observaciones = data.get("observaciones", "Sin observaciones").split('\n')
        for i, obs in enumerate(observaciones, start=3):
            ws2[f'A{i}'] = obs
        
        # Agregar hoja de información del sistema
        ws3 = wb.create_sheet(title="Información Sistema")
        info_sistema = [
            ["INFORMACIÓN DEL SISTEMA"],
            ["Procesado por:", data.get("procesado_por", "Sistema OCR")],
            ["Fecha procesamiento:", data.get("fecha_procesamiento", "N/A")],
            ["Modo OCR:", data.get("modo_ocr", "SIMULADO")],
            ["Nombre archivo:", os.path.basename(output_path).replace('.xlsx', '.pdf')],
            ["Generado el:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
        
        for row in info_sistema:
            ws3.append(row)
        
        # Guardar archivo
        wb.save(output_path)
        logger.info(f"Excel generado: {output_path}")
        
        return output_path
        
    except Exception as e:
        logger.error(f"Error generando Excel: {e}")
        # Crear un Excel mínimo si falla
        return generate_simple_excel(data, output_path)


def generate_simple_excel(data: dict, output_path: str) -> str:
    """Genera un Excel simple como fallback"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        
        # Datos básicos
        ws.append(["CAMPO", "VALOR"])
        ws.append(["Número Proceso", data.get("process_number", "")])
        ws.append(["Proveedor", data.get("provider", "")])
        ws.append(["Conductor", data.get("driver", "")])
        ws.append(["Placa", data.get("plate_tractor", "")])
        ws.append(["Peso Neto", f"{data.get('net_weight', 0):,.2f} kg"])
        ws.append(["Producto", data.get("product", "")])
        ws.append(["Fecha", data.get("fecha", "")])
        
        wb.save(output_path)
        logger.info(f"Excel simple generado: {output_path}")
        return output_path
        
    except Exception as e:
        logger.error(f"Error generando Excel simple: {e}")
        raise


# Necesario para datetime
from datetime import datetime
