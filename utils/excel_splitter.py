# utils/excel_splitter.py
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import re
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook, Workbook


@dataclass
class SplitConfig:
    salida_base: Path
    anio: int = 2025
    mes_inicio: int = 4
    mes_fin: int = 12


def _norm(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\n", " ").replace("\r", " ")
    s = s.replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace('"', "").replace("“", "").replace("”", "")
    return s


def _parse_sheet_date(sheet_name: str) -> Optional[datetime]:
    name = _norm(sheet_name)
    try:
        return datetime.strptime(name, "%d-%m-%Y")
    except Exception:
        return None


def _find_header_row_and_map(ws) -> Tuple[int, Dict[str, int]]:
    """
    Busca la fila header (primeras 30 filas) y devuelve:
    - row_idx (1-based)
    - mapping: nombre_col_estandar -> indice_col (0-based)
    """
    required = {
        "Item": ["Item"],
        "Código del Material": ["Código del Material", "Codigo del Material", "Codigo", "COD", "Material"],
        "Texto breve de material": ["Texto breve de material", "Texto breve", "Descripcion", "Descripción", "Texto"],
        "Unidad Medida": ["Unidad Medida", "Unidad", "Unidad de medida", "Unidad de medida base", "U.M.", "UM"],
        "Ubicación": ["Ubicación", "Ubicacion", "Location", "UBI"],
        "Fisico": ["Fisico", "Físico", "Libre utilización", "Libre utilizacion", "Cantidad", "Stock contado"],
        "STOCK": ["STOCK", "Stock", "Stock sistema", "SISTEMA"],
        "Difere": ["Difere", "Difer", "Diferencia"],
        "Observac.": ["Observac.", "Observacion", "Observación", "Obs", "Observaciones"],
    }

    # normalizamos los alias
    required_norm = {k: [_norm(x).lower() for x in v] for k, v in required.items()}

    best_row = None
    best_map = {}

    for r in range(1, 31):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        headers = [_norm(v).lower() for v in row_vals]

        mapping: Dict[str, int] = {}
        for std_name, aliases in required_norm.items():
            for idx, h in enumerate(headers):
                if h in aliases:
                    mapping[std_name] = idx
                    break

        # Si encontramos al menos estas 6, consideramos que es header válido
        min_core = ["Código del Material", "Texto breve de material", "Unidad Medida", "Ubicación", "Fisico"]
        score = sum(1 for x in min_core if x in mapping)

        if score >= 4:
            # guardamos el mejor
            if best_row is None or len(mapping) > len(best_map):
                best_row = r
                best_map = mapping

            # si ya encontramos casi todo, salimos
            if len(mapping) >= 7:
                break

    if best_row is None:
        raise Exception("No se encontró fila de encabezados (header) en las primeras 30 filas.")

    # Validación mínima: las que tú dijiste necesarias
    necesarios = ["Item", "Código del Material", "Texto breve de material", "Unidad Medida", "Ubicación", "Fisico", "STOCK", "Difere", "Observac."]
    faltantes = [c for c in necesarios if c not in best_map]
    if faltantes:
        raise Exception(f"❌ Columnas faltantes: {faltantes}")

    return best_row, best_map


def dividir_excel_por_dias(
    archivo_excel: str | Path,
    salida_base: str | Path = "inventarios_procesados",
    anio: int = 2025,
    mes_inicio: int = 4,
    mes_fin: int = 12,
) -> List[Path]:
    """
    Divide un Excel con muchas hojas (cada hoja = día) en archivos diarios.
    Retorna lista de paths generados.
    """
    archivo_excel = Path(archivo_excel)
    if not archivo_excel.exists():
        raise Exception(f"Archivo no existe: {archivo_excel}")

    cfg = SplitConfig(salida_base=Path(salida_base), anio=anio, mes_inicio=mes_inicio, mes_fin=mes_fin)
    cfg.salida_base.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=str(archivo_excel), read_only=True, data_only=True)
    generados: List[Path] = []

    for sheet_name in wb.sheetnames:
        fecha = _parse_sheet_date(sheet_name)
        if not fecha:
            continue

        if fecha.year != cfg.anio:
            continue
        if not (cfg.mes_inicio <= fecha.month <= cfg.mes_fin):
            continue

        ws = wb[sheet_name]

        header_row, colmap = _find_header_row_and_map(ws)

        # destino: inventarios_procesados/2025/04/inventario_2025_04_10.xlsx
        out_dir = cfg.salida_base / f"{fecha.year}" / f"{fecha.month:02d}"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"inventario_{fecha:%Y_%m_%d}.xlsx"

        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = f"{fecha:%d-%m-%Y}"

        # Header EXACTO (lo que tú necesitas)
        headers_out = ["Item", "Código del Material", "Texto breve de material", "Unidad Medida", "Ubicación", "Fisico", "STOCK", "Difere", "Observac."]
        out_ws.append(headers_out)

        # Recorremos filas desde la siguiente al header
        max_r = ws.max_row
        for r in range(header_row + 1, max_r + 1):
            row = []
            # extraemos en el orden de salida
            for h in headers_out:
                idx0 = colmap[h]  # 0-based
                val = ws.cell(row=r, column=idx0 + 1).value

                if h in ("Ubicación",):
                    val = _norm(val).replace(" ", "").upper()
                else:
                    val = _norm(val)

                row.append(val)

            # si no hay código de material, saltamos
            if not row[1]:
                continue

            out_ws.append(row)

        out_wb.save(str(out_path))
        generados.append(out_path)

    wb.close()

    if not generados:
        raise Exception("No se generó ningún archivo. Revisa nombres de hojas (dd-mm-YYYY) y rango Abril–Diciembre.")

    return generados
