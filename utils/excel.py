from io import BytesIO
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# INVENTARIO BASE
# =============================================================================
def load_inventory_excel(file):
    df = pd.read_excel(file)

    columnas = [
        "Código del Material",
        "Texto breve de material",
        "Unidad de medida base",
        "Ubicación",
        "Libre utilización",
    ]

    for c in columnas:
        if c not in df.columns:
            raise Exception(f"❌ Falta columna: {c}")

    df = df.copy()
    df["Código del Material"] = df["Código del Material"].astype(str).str.strip()
    df["Texto breve de material"] = df["Texto breve de material"].astype(str).str.strip()
    df["Unidad de medida base"] = df["Unidad de medida base"].astype(str).str.strip()
    df["Ubicación"] = df["Ubicación"].astype(str).str.replace(" ", "").str.upper().str.strip()
    df["Libre utilización"] = pd.to_numeric(df["Libre utilización"], errors="coerce").fillna(0)

    return df

# =============================================================================
# INVENTARIO HISTÓRICO (EXCEL ANTIGUO)
# =============================================================================
def load_inventory_historic_excel(file):
    df = pd.read_excel(file)

    columnas_requeridas = [
        "Código del Material",
        "Texto breve de material",
        "Unidad Medida",
        "Ubicación",
        "Fisico",
        "STOCK",
        "Difere",
        "Observac.",
    ]

    for c in columnas_requeridas:
        if c not in df.columns:
            raise Exception(f"❌ Falta columna histórica: {c}")

    df = df.copy()

    df["Código del Material"] = df["Código del Material"].astype(str).str.strip()
    df["Texto breve de material"] = df["Texto breve de material"].astype(str).str.strip()
    df["Unidad Medida"] = df["Unidad Medida"].astype(str).str.strip()

    df["Ubicación"] = (
        df["Ubicación"]
        .astype(str)
        .str.replace(" ", "")
        .str.upper()
        .str.strip()
    )

    df["Fisico"] = pd.to_numeric(df["Fisico"], errors="coerce").fillna(0)
    df["STOCK"] = pd.to_numeric(df["STOCK"], errors="coerce").fillna(0)
    df["Difere"] = pd.to_numeric(df["Difere"], errors="coerce").fillna(0)
    df["Observac."] = df["Observac."].astype(str).fillna("")

    return df
# =============================================================================
# ALMACÉN 2D
# =============================================================================
def load_warehouse2d_excel(file):
    df = pd.read_excel(file)

    columnas_requeridas = [
        "Código del Material",
        "Texto breve de material",
        "Unidad de medida base",
        "Stock de seguridad",
        "Stock máximo",
        "Ubicación",
        "Libre utilización",
    ]

    for c in columnas_requeridas:
        if c not in df.columns:
            raise Exception(f"❌ Falta columna obligatoria en almacén 2D: {c}")

    df = df.copy()

    df["Código del Material"] = df["Código del Material"].astype(str).str.strip()
    df["Texto breve de material"] = df["Texto breve de material"].astype(str).str.strip()
    df["Unidad de medida base"] = df["Unidad de medida base"].astype(str).str.strip()

    df["Ubicación"] = (
        df["Ubicación"]
        .astype(str)
        .str.strip()
        .str.replace(" ", "")
        .str.upper()
    )

    df["Stock de seguridad"] = pd.to_numeric(df["Stock de seguridad"], errors="coerce").fillna(0)
    df["Stock máximo"] = pd.to_numeric(df["Stock máximo"], errors="coerce").fillna(0)
    df["Libre utilización"] = pd.to_numeric(df["Libre utilización"], errors="coerce").fillna(0)

    return df

# =============================================================================
# ORDENAR UBICACIONES
# =============================================================================
def sort_location_advanced(loc):
    try:
        loc = str(loc).replace(" ", "").upper()
        if loc.startswith("E"):
            nums = "".join(c for c in loc if c.isdigit())
            return int(nums) if nums else 999999
        return 999999
    except:
        return 999999

# =============================================================================
# EXCEL CON FÓRMULA DE ESTADO DIRECTAMENTE EN LAS CELDAS
# =============================================================================
def generate_discrepancies_excel(df, meta=None):

    output = BytesIO()
    wb = Workbook()

    # =========================
    # Validación
    # =========================
    if df is None or df.empty:
        ws = wb.active
        ws.title = "DISCREPANCIAS"
        ws.append(["SIN DATOS PARA EXPORTAR"])
        wb.save(output)
        output.seek(0)
        return output

    df = df.copy()

    # =========================
    # Columnas estándar
    # =========================
    columnas = [
        "Código Material",
        "Descripción", 
        "Unidad",
        "Ubicación",
        "Stock sistema",
        "Stock contado",
        "Diferencia",
        "Estado",
    ]

    # Asegurar que existan todas las columnas
    for c in columnas:
        if c not in df.columns:
            if "Stock" in c or c == "Diferencia":
                df[c] = 0
            elif c == "Estado":
                df[c] = ""

    # Asegurar que los valores sean numéricos
    df["Stock sistema"] = pd.to_numeric(df["Stock sistema"], errors='coerce').fillna(0)
    df["Stock contado"] = pd.to_numeric(df["Stock contado"], errors='coerce').fillna(0)
    
    # Calcular diferencia REAL
    df["Diferencia"] = df["Stock contado"] - df["Stock sistema"]
    
    # Calcular estado REAL (para mostrar inicialmente)
    def estado(row):
        stock_contado = row.get("Stock contado", 0)
        diferencia = row.get("Diferencia", 0)
        
        if stock_contado == 0:
            return "NO CONTADO"
        if diferencia == 0:
            return "OK"
        if diferencia < 0:
            return "FALTA" if abs(diferencia) < 5 else "CRÍTICO"
        return "SOBRA"

    df["Estado"] = df.apply(estado, axis=1)

    # =========================
    # ESTILOS
    # =========================
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    bold = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # COLORES
    fill_ok = PatternFill("solid", fgColor="C6EFCE")
    fill_falta = PatternFill("solid", fgColor="FFEB9C")
    fill_critico = PatternFill("solid", fgColor="FFC7CE")
    fill_sobra = PatternFill("solid", fgColor="FFEB9C")
    fill_nocont = PatternFill("solid", fgColor="E7E6E6")

    # =========================
    # HOJA 1 – RESUMEN
    # =========================
    ws0 = wb.active
    ws0.title = "RESUMEN"

    ws0["A1"] = "REPORTE DE DISCREPANCIAS – WAREHOUSE MRO"
    ws0.merge_cells("A1:F1")
    ws0["A1"].font = title_font
    ws0["A1"].fill = header_fill
    ws0["A1"].alignment = left

    generado_por = (meta or {}).get("generado_por", "Sistema MRO")
    generado_en = (meta or {}).get(
        "generado_en",
        datetime.now().strftime("%Y-%m-%d %H:%M")
    )

    ws0["A3"] = "Generado por:"
    ws0["B3"] = generado_por
    ws0["A4"] = "Generado en:"
    ws0["B4"] = generado_en

    for row in range(3, 5):
        ws0[f"A{row}"].font = bold

    total = len(df)
    ok = (df["Estado"] == "OK").sum()
    exactitud = round((ok / total) * 100, 2) if total else 0

    ws0["D3"] = "Total ítems:"
    ws0["E3"] = total
    ws0["D4"] = "Exactitud:"
    ws0["E4"] = f"{exactitud}%"

    for row in range(3, 5):
        ws0[f"D{row}"].font = bold

    ws0["A6"] = "Resumen por Estado"
    ws0["A6"].font = Font(bold=True, size=12)

    ws0.append(["Estado", "Cantidad"])
    ws0["A7"].font = bold
    ws0["B7"].font = bold

    estados = ["OK", "FALTA", "CRÍTICO", "SOBRA", "NO CONTADO"]
    fila = 8
    for estado_lbl in estados:
        ws0[f"A{fila}"] = estado_lbl
        ws0[f"B{fila}"] = int((df["Estado"] == estado_lbl).sum())
        fila += 1

    for col in range(1, 7):
        ws0.column_dimensions[get_column_letter(col)].width = 22

    # =========================
    # HOJA 2 – DETALLE CON FÓRMULAS REALES
    # =========================
    ws = wb.create_sheet("DISCREPANCIAS")
    
    # Escribir encabezados
    ws.append(columnas)

    # Escribir datos CON FÓRMULAS REALES
    for i, row in df.iterrows():
        row_num = i + 2  # +2 porque la fila 1 es encabezado
        
        # FÓRMULA CORREGIDA - SIN PROBLEMAS DE CARACTERES ESPECIALES
        formula_estado = f'IF(F{row_num}=0,"NO CONTADO",IF(G{row_num}=0,"OK",IF(G{row_num}<0,IF(ABS(G{row_num})<5,"FALTA","CRÍTICO"),"SOBRA")))'
        
        ws.append([
            row["Código Material"],
            row["Descripción"],
            row["Unidad"],
            row["Ubicación"],
            row["Stock sistema"],  # Valor
            row["Stock contado"],   # Valor
            f'F{row_num}-E{row_num}',  # Fórmula diferencia
            formula_estado,  # Fórmula estado
        ])

    # Estilos para encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Ajustar anchos
    widths = [20, 45, 10, 14, 16, 16, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Congelar paneles y filtros
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # =========================
    # APLICAR COLORES SEGÚN FÓRMULA DE ESTADO
    # =========================
    from openpyxl.formatting.rule import FormulaRule
    
    # Aplicar formatos condicionales basados en la fórmula de estado
    # VERDE para OK
    ok_rule = FormulaRule(formula=[f'$H2="OK"'], fill=fill_ok)
    ws.conditional_formatting.add(f'A2:H{total+1}', ok_rule)
    
    # AMARILLO para FALTA
    falta_rule = FormulaRule(formula=[f'$H2="FALTA"'], fill=fill_falta)
    ws.conditional_formatting.add(f'A2:H{total+1}', falta_rule)
    
    # ROJO para CRÍTICO
    critico_rule = FormulaRule(formula=[f'$H2="CRÍTICO"'], fill=fill_critico)
    ws.conditional_formatting.add(f'A2:H{total+1}', critico_rule)
    
    # AMARILLO para SOBRA
    sobra_rule = FormulaRule(formula=[f'$H2="SOBRA"'], fill=fill_sobra)
    ws.conditional_formatting.add(f'A2:H{total+1}', sobra_rule)
    
    # GRIS para NO CONTADO
    nocont_rule = FormulaRule(formula=[f'$H2="NO CONTADO"'], fill=fill_nocont)
    ws.conditional_formatting.add(f'A2:H{total+1}', nocont_rule)

    # Aplicar bordes y alineaciones
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = border
            if cell.column in [5, 6, 7]:  # Números
                cell.alignment = right
                if cell.column == 7:  # Diferencia
                    cell.number_format = '0'
            elif cell.column in [1, 3, 4, 8]:  # Centrado
                cell.alignment = center
            else:  # Izquierda
                cell.alignment = left

    # =========================
    # HOJA 3 - INSTRUCCIONES
    # =========================
    ws2 = wb.create_sheet("INSTRUCCIONES")
    
    ws2["A1"] = "CÓMO USAR ESTE EXCEL"
    ws2.merge_cells("A1:D1")
    ws2["A1"].font = title_font
    ws2["A1"].fill = header_fill
    ws2["A1"].alignment = center

    instrucciones = [
        ("¡LISTO!", "Las fórmulas ya están incluidas", "", ""),
        ("", "", "", ""),
        ("COLUMNA G (Diferencia):", "Fórmula: =F2-E2", "Se calcula automáticamente", ""),
        ("COLUMNA H (Estado):", 'Fórmula: =IF(F2=0,"NO CONTADO",IF(G2=0,"OK",IF(G2<0,IF(ABS(G2)<5,"FALTA","CRÍTICO"),"SOBRA")))', "Se calcula automáticamente", ""),
        ("", "", "", ""),
        ("QUÉ MODIFICAR:", "", "", ""),
        ("1", "Columna E: Stock sistema", "Modifica valores según necesites", ""),
        ("2", "Columna F: Stock contado", "Modifica valores según inventario real", ""),
        ("", "", "", ""),
        ("QUÉ PASA AUTOMÁTICAMENTE:", "", "", ""),
        ("•", "Columna G se recalcula", "= Contado - Sistema", ""),
        ("•", "Columna H se actualiza", "Según las reglas definidas", ""),
        ("•", "Color de fila cambia", "Según el estado resultante", ""),
        ("", "", "", ""),
        ("EJEMPLO:", "", "", ""),
        ("Si E2=10 y F2=8:", "G2 = 8-10 = -2", "H2 = FALTA", "Color AMARILLO"),
        ("Si cambias F2 a 12:", "G2 = 12-10 = 2", "H2 = SOBRA", "Color AMARILLO"),
        ("Si cambias F2 a 5:", "G2 = 5-10 = -5", "H2 = CRÍTICO", "Color ROJO"),
        ("Si F2=0:", "G2 = 0-10 = -10", "H2 = NO CONTADO", "Color GRIS"),
    ]

    fila_inst = 3
    for inst in instrucciones:
        ws2.append(inst)
        if fila_inst == 3:
            for col in range(1, 5):
                ws2.cell(row=fila_inst, column=col).font = bold
                ws2.cell(row=fila_inst, column=col).fill = PatternFill("solid", fgColor="E0E0E0")
        fila_inst += 1

    # Ajustar anchos
    for col, width in zip(["A", "B", "C", "D"], [15, 40, 30, 20]):
        ws2.column_dimensions[col].width = width

    # =========================
    # GUARDAR
    # =========================
    wb.save(output)
    output.seek(0)
    return output
# =============================================================================
# EXPORTAR SNAPSHOT HISTÓRICO A EXCEL  ✅ (ESTO FALTABA)
# =============================================================================
def generate_history_snapshot_excel(items, snapshot_name):

    wb = Workbook()
    ws = wb.active
    ws.title = "INVENTARIO"

    headers = [
        "Código Material",
        "Descripción",
        "Unidad",
        "Ubicación",
        "Stock",
        "Fecha",
    ]
    ws.append(headers)

    for i in items:
        ws.append([
            i.material_code,
            i.material_text,
            i.base_unit,
            i.location,
            i.libre_utilizacion,
            i.creado_en.strftime("%d/%m/%Y") if i.creado_en else "",
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output






